import openpyxl as xl
import pprint
import pandas as pd
import numpy as np

file=input('Enter a file: ') # also call the second file 
print('Evoking the file....')
file1=xl.load_workbook(file) # load the second file as file2
print('loading the file....')

# sheet1title = input ('enter sheet 1 title: ') 
# sheet1 = file1[sheet1title]
# sheet2title = input ('enter sheet 2 title: ') # call the sheet1 of the second file 
# sheet2 = file1[sheet2title] # sheet2 now becomes sheet1 of the second file ie file2[sheet1title]
# print('reading workbook......')

# sr_sales_A_S = [] # open an empty list
# for row in range(2, sheet2.max_row + 1): # Identify the rows to work with in the first sheet
#     Meter_Acount_Position = sheet2['C' + str(row)].value #{*now becomes sheet1*} reading ONLY the column of interest
#     sr_sales_A_S.append(Meter_Acount_Position) # Attaching the column of interest above to an empty list above
    
print('reading rows...')

for rowNum in range(2, sheet1.max_row + 1): # Identify the rows to compare with in the second sheet
    Meter_Acount_Number_Common = sheet1.cell(row=rowNum, column=3).value  # equating the columns of interest of both sheets by connecting sheet1/sheet2 via their row(index)
    if Meter_Acount_Number_Common in sr_sales_A_S:   # conditional statement
        sheet1.cell(row=rowNum, column=7).value = 'yes' # where to store the result_yes
        
print('rounding up......')
file1.save('Ok2.xlsx')
print('Finally_Saved')
file1.close()

# ......continue please