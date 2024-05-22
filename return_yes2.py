
import openpyxl as xl
import pprint
#print('enter document name...')
docname = input('enter document name with extension:')
print('Opening workbook...')
file1 = xl.load_workbook(docname)
sheet1title = input ('enter sheet 1 title: ')
sheet1 = file1[sheet1title]
sheet2title = input ('enter sheet 2 title: ')
sheet2 = file1[sheet2title]
sheet3title = input ('enter sheet 3 title: ')
sheet3 = file1[sheet3title]
print('reading workbook')

sr_sales = []
print('Reading rows...')

for row1 in range(2, sheet2.max_row + 1):
    agent = sheet2['A' +str(row1)].value
    funding = sheet2['I' + str(row1)].value
    sr_sales.append(funding)

for row2 in range(2, sheet3.max_row + 1):
    agent = sheet3['A' +str(row2)].value
    funding = sheet3['D' + str(row2)].value
    sr_sales.append(funding)

    #sr_sales[agent]['funding'] = funding
    #See how to roll out another comparison sheet

print('Reading rows...')
for rowNum1 in range(2, sheet1.max_row + 1):
    agent_a1 = sheet1.cell(row=rowNum1, column=4).value
    if agent_a1 in sr_sales:
        sheet1.cell(row=rowNum1, column=5).value = 'yes1'

for rowNum2 in range(2, sheet1.max_row + 1):
    agent_a2 = sheet1.cell(row=rowNum2, column=4).value
    if agent_a2 in sr_sales:
        sheet1.cell(row=rowNum2, column=5).value = 'yes2'

print ('finishing up...saving...') 
file1.save('new_file.xlsx')
file1.close()