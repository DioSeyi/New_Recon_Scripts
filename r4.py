import openpyxl as xl
import pprint
import pandas as pd
import numpy as np

file=input('Enter a file: ')  
print('Evoking the file....')
file1=xl.load_workbook(file)
print('loading the file....')

sheet1title=input ('enter sheet 1 title: ') 
sheet1=file1[sheet1title]
print('reading workbook......')

sr_sales_A_S = []
for row in range(2, sheet1.max_row + 1): 
    Meter_Acount_Position = sheet1['E' + str(row)].value
    sr_sales_A_S.append(Meter_Acount_Position) 
    
print('reading rows...')
 
for rowNum1 in range(2, sheet1.max_row + 1): 
    Meter_Acount_Number_Common = sheet1.cell(row=rowNum1, column = 19).value
    if Meter_Acount_Number_Common in sr_sales_A_S:
        sheet1.cell(row=rowNum1, column = 14).value = 'yes'
        
for rowNum2 in range(2, sheet1.max_row + 1): 
    Meter_Acount_Number_Common = sheet1.cell(row=rowNum2, column = 20).value
    if Meter_Acount_Number_Common in sr_sales_A_S:
        sheet1.cell(row=rowNum2, column = 15).value = 'yes'
        
print('rounding up......')
file1.save('Ok4.xlsx')

print('Finally_Saved')
file1.close()

# ......continue please Passed