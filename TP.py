#!/usr/bin/env python
# coding: utf-8

# In[3]:


cd C:\Users\ONE CARD\Desktop\pydata


# In[4]:


pwd!


# In[5]:


get_ipython().system('pip install xlsxwriter')


# In[6]:


import os
import pandas as pd
import numpy as np
import datetime as dt
from glob import glob
import xlsxwriter
from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment,fills
pd.set_option("display.max_rows", 5)


# In[26]:


dm = pd.read_excel('1_2_31.xlsx')


# In[27]:


dm.head(3)


# In[28]:


dm.columns


# In[29]:


dm1 = dm['User Parent'].str.contains('Vendemitria Limited',na = False)|dm['User Parent'].str.contains('Yahaya Eniola Fausat',na = False)|dm['User Parent'].str.contains('Sylvester Emeka Ugbala',na = False)|dm['User Parent'].str.contains('OKWELOGU CHINAZO',na = False)|dm['User Parent'].str.contains('OKWELOGU CHINAZO',na = False)|dm['User'].str.contains('Yahaya Eniola Fausat',na = False)|dm['User Parent'].str.contains('Nwaefulu Emmanuel Emeka',na = False)# add Emeka |dm['Transaction Status'].str.contains('Success',na = False)
dm1 = dm.loc[dm1]
# dm2 = (dm1['Transaction Status'] == 'Success') why cant it remove the failures here but there
# dm2 = dm1.loc[dm2]


# In[30]:


dm1.head(3)


# In[31]:


dm1.to_excel('dm1.xlsx',index=False)


# In[32]:


dm2 = pd.read_excel('dm1.xlsx')


# In[33]:


dm2 = (dm1['Transaction Status'] == 'Success')
dm2 = dm1.loc[dm2]


# In[34]:


dm2.to_excel('dm2.xlsx',index=False)


# In[ ]:




