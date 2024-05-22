import openpyxl as xl
import pprint
#print('enter document name...')
docname = input('enter document name with extension:')
print('Opening workbook...')
file1 = xl.load_workbook(docname)
sheet1title = input ('enter sheet 1 title: ')
sheet1 = file1[sheet1title]
sheet2title = input ('enter sheet 2 title: ')
sheet2 = file1[sheet2title]
print('reading workbook')

sr_sales = {}
print('Reading rows...')
for row in range(2, sheet2.max_row + 1):
	agent = sheet2['E' +str(row)].value
	funding = sheet2['F' + str(row)].value
	sr_sales.setdefault(agent, {'funding':0})
	#osadata.setdefault({})
	#osadata.setdefault()
	sr_sales[agent]['funding'] += int(funding)
	
print('Reading rows...')
for rowNum in range(2, sheet1.max_row + 1):
	agent_a = sheet1.cell(row=rowNum, column=3).value
	if agent_a in sr_sales:
		sheet1.cell(row=rowNum, column=4).value = sr_sales[agent_a]['funding']
print ('finishing up...saving...') 
file1.save('DTU_SR_Sales_16-15_Jan.xlsx')
