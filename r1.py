import openpyxl as xl
import pprint
import pandas as pd
import numpy as np

file_1 = input('Enter a file1: ') # also call the second file
file_2 = input('Enter a file2: ')
print('Evoking the files....')
file1=xl.load_workbook(file_1)
file2=xl.load_workbook(file_2) # load the second file as file2
print('loading the file....')

sheet1title = input ('enter 1_sheet_wk1 title: ') 
sheet1 = file1[sheet1title]
sheet2title = input ('enter 1_sheet_wk2 title: ')   # call the sheet1 of the second file 
sheet2 = file2[sheet2title]    # sheet2 now becomes sheet1 of the second file ie file2[sheet1title]
print('reading workbook......')

sr_sales_A_S = [] # open an empty list
for row in range(2, sheet2.max_row + 1): # Identify the rows to work with in the first sheet
    Meter_Acount_Position = sheet2['C' + str(row)].value #{*now becomes sheet1*} reading ONLY the column of interest
    sr_sales_A_S.append(Meter_Acount_Position) # Attaching the column of interest above to an empty list above
    
print('reading rows...')

for rowNum in range(2, sheet1.max_row + 1): # Identify the rows to compare with in the second sheet
    Meter_Acount_Number_Common = sheet1.cell(row=rowNum, column = 3).value  # equating the columns of interest of both sheets by connecting sheet1/sheet2 via their row(index)
    if Meter_Acount_Number_Common in sr_sales_A_S:
        sheet1.cell(row = rowNum, column = 5).value = 'yes'
        
print('rounding up......')
file1.save('Ok1.xlsx')
print('Finally_Saved')
file1.close()

# ......continue please  