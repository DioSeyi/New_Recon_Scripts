import openpyxl as xl
import pprint
import pandas as pd
import numpy as np

file=input('Enter a file: ')  
print('Evoking the file....')
file1=xl.load_workbook(file)
print('loading the file....')

sheet1title=input ('enter sheet 1 title: ') 
sheet1=file1[sheet1title]
print('reading workbook......')

sr_sales_A_S = []
for row in range(2, sheet1.max_row + 1): 
    Meter_Acount_Position = sheet1['C' + str(row)].value  # The column can be change to column of interest depending on which to compare
    sr_sales_A_S.append(Meter_Acount_Position) 
    
print('reading rows...')

for rowNum in range(2, sheet1.max_row + 1): 
    Meter_Acount_Number_Common = sheet1.cell(row=rowNum, column = 9).value  #  column same as here to 
    if Meter_Acount_Number_Common in sr_sales_A_S:
        sheet1.cell(row=rowNum, column = 21).value = 'yes'  #  column same as here to 
        
print('rounding up......')
file1.save('Ok3.xlsx')
print('Finally_Saved')
file1.close()

# ......continue please  Passed