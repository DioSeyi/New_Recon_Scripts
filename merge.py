import os
import pandas as pd
import numpy as np
import datetime as dt
from glob import glob
import xlsxwriter
from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment,fills
pd.set_option("display.max_rows", 5)


tab_1 = glob('transactiondetail_*.csv')
tab_2 = pd.concat([pd.read_csv(f) for f in tab_1], sort = False) # List comprehension
tab_2.drop(['Transaction Detail','Unnamed: 3','Unnamed: 4','Unnamed: 9'],axis=1,inplace=True)
tab_2.columns = tab_2.iloc[2]
tab_2.drop([0,1,2],inplace = True)
tab_9 = tab_2['Result Description'].str.contains('Transaction Successful',na = False)|tab_2['Result Description'].str.contains('InProgress',na = False)
tab_9 = tab_2.loc[tab_9]
tab_10 = (tab_9['Trans Type'] == 'Topup') | (tab_9['Trans Type'] == 'Fund Transfer') | (tab_9['Trans Type'] == 'Wallet Adjustment') | (tab_9['Trans Type'] == 'Wallet Transfer') | (tab_9['Trans Type'] == 'Wallet Topup') | (tab_9['Trans Type'] == 'PINREDEEM') # To disregards any other top inclusive like 'Wallet Topup'
tab_10 = tab_9.loc[tab_10]
tab_10['Client Type'].fillna('B2B',inplace=True)
tab_10.to_csv('tab_10.csv')

# # For Putting Report Together

tab_1 = glob('transactiondetail_*.csv')  
tab_3 = pd.concat([pd.read_csv(f) for f in tab_1] ,sort = False) # List comprehension
tab_3.drop(['Transaction Detail','Unnamed: 3','Unnamed: 4'],axis=1,inplace=True)
tab_3.columns = tab_3.iloc[2]
tab_4 = tab_3['Result Description'].str.contains('Transaction Successful',na = False)|tab_3['Result Description'].str.contains('InProgress',na = False)
tab_4 = tab_3.loc[tab_4]
tab_5 = (tab_4['Trans Type'] == 'Topup') | (tab_4['Trans Type'] == 'Fund Transfer') # To disregards any other top inclusive like 'Wallet Topup'
tab_5 = tab_4.loc[tab_5]
tab_5.to_csv('tab_5.csv',index = False)

data_k2 = pd.read_csv('tab_5.csv')

# #.......IBEDC Report
data_k3 = (data_k2['Operator'] == 'IBEDC')
data_Ibedc = data_k2.loc[data_k3]
data_Ibedc.to_csv('IBEDC.csv',index = False)

# #.......KANDOELEC Report
data_k3 = (data_k2['Operator'] == 'KANDOELEC')
data_k3 = data_k2.loc[data_k3]
data_k4 = (data_k3['Agent Code'] == '04001010001')
data_k5 = (data_k3['Agent Code'] != '04001010001')
data_k6 = data_Kandoelec_SR = data_k3.loc[data_k4]
data_k7 = data_Kandoelec_TP = data_k3.loc[data_k5]
data_k6.to_csv('KANDOELECSRs1.csv',index = False)
data_k7.to_csv('KANDOELECTPs2.csv',index = False)

# #.......EKOELEC Report
data_k3 = (data_k2['Operator'] == 'EkoElec')
data_k3 = data_k2.loc[data_k3]
data_k4 = (data_k3['Agent Code'] == '03005000004')|(data_k3['Agent Code'] == '03005000005')|(data_k3['Agent Code'] == '03005000006')
data_k5 = (data_k3['Agent Code'] != '03005000004') & (data_k3['Agent Code'] != '03005000005') & (data_k3['Agent Code'] != '03005000006')
data_k8 = data_Ekepd_SR = data_k3.loc[data_k4]
data_k9 = data_Ekedp_TP = data_k3.loc[data_k5]
data_k8.to_csv('EKEDPSRs1.csv',index = False)
data_k9.to_csv('EKEDPTPs2.csv',index = False)

# #.......JEDC Report
# data_k3 = (data_k2['Operator'] == 'JEDC')
# data_k3 = data_k2.loc[data_k3]
# data_k4 = (data_k3['Agent Code'] == '03001010002') | (data_k3['Agent Code'] == '03001010003') | (data_k3['Agent Code'] == '03001010004') | (data_k3['Agent Code'] == '03001010005') | (data_k3['Agent Code'] == '03001010006') | (data_k3['Agent Code'] == '03001010007') | (data_k3['Agent Code'] == '03001010009') | (data_k3['Agent Code'] == '03001010011') | (data_k3['Agent Code'] == '03001010012') | (data_k3['Agent Code'] == '03001010013') | (data_k3['Agent Code'] == '03001020001') | (data_k3['Agent Code'] == '03001020002') | (data_k3['Agent Code'] == '03001020003') | (data_k3['Agent Code'] == '03001020004') | (data_k3['Agent Code'] == '03001020005') | (data_k3['Agent Code'] == '03001020007') | (data_k3['Agent Code'] == '03001020008') | (data_k3['Agent Code'] == '03001020012') | (data_k3['Agent Code'] == '03001020013') | (data_k3['Agent Code'] == '03001030001') | (data_k3['Agent Code'] == '03001030002') | (data_k3['Agent Code'] == '03001030003') | (data_k3['Agent Code'] == '03001030004') | (data_k3['Agent Code'] == '03001030005') | (data_k3['Agent Code'] == '03001030006') | (data_k3['Agent Code'] == '03001030007') | (data_k3['Agent Code'] == '03001030008') | (data_k3['Agent Code'] == '03001030009') | (data_k3['Agent Code'] == '03001040001') | (data_k3['Agent Code'] == '03001040002') | (data_k3['Agent Code'] == '03001040003') | (data_k3['Agent Code'] == '03001040004') | (data_k3['Agent Code'] == '03001040005') | (data_k3['Agent Code'] == '03001040006') | (data_k3['Agent Code'] == '03001040007')
# data_k5 = (data_k3['Agent Code'] != '03001010002') & (data_k3['Agent Code'] != '03001010003') & (data_k3['Agent Code'] != '03001010004') & (data_k3['Agent Code'] != '03001010005') & (data_k3['Agent Code'] != '03001010006') & (data_k3['Agent Code'] != '03001010007') & (data_k3['Agent Code'] != '03001010009') & (data_k3['Agent Code'] != '03001010011') & (data_k3['Agent Code'] != '03001010012') & (data_k3['Agent Code'] != '03001010013') & (data_k3['Agent Code'] != '03001020001') & (data_k3['Agent Code'] != '03001020002') & (data_k3['Agent Code'] != '03001020003') & (data_k3['Agent Code'] != '03001020004') & (data_k3['Agent Code'] != '03001020005') & (data_k3['Agent Code'] != '03001020007') & (data_k3['Agent Code'] != '03001020008') & (data_k3['Agent Code'] != '03001020012') & (data_k3['Agent Code'] != '03001020013') & (data_k3['Agent Code'] != '03001030001') & (data_k3['Agent Code'] != '03001030002') & (data_k3['Agent Code'] != '03001030003') & (data_k3['Agent Code'] != '03001030004') & (data_k3['Agent Code'] != '03001030005') & (data_k3['Agent Code'] != '03001030006') & (data_k3['Agent Code'] != '03001030007') & (data_k3['Agent Code'] != '03001030008') & (data_k3['Agent Code'] != '03001030009') & (data_k3['Agent Code'] != '03001040001') & (data_k3['Agent Code'] != '03001040002') & (data_k3['Agent Code'] != '03001040003') & (data_k3['Agent Code'] != '03001040004') & (data_k3['Agent Code'] != '03001040005') & (data_k3['Agent Code'] != '03001040006') & (data_k3['Agent Code'] != '03001040007')
# data_Jedc_SR = data_k3.loc[data_k4]
# data_Jedc_TP = data_k3.loc[data_k5]
# data_k6.to_csv('JEDCSRs.csv',index = False)
# data_k7.to_csv('JEDCTPs.csv',index = False)

# #.......All Report sheets in a single workbook
# writer = pd.ExcelWriter(' date OneCard TPs and SRs Electricity Utilization Report Dec 2021checking.xlsx')
# data_Ibedc.to_excel(writer, sheet_name = 'IBEDC', index = False)
# data_Kandoelec_SR.to_excel(writer, sheet_name = 'KANDOELECSRs', index = False)
# data_Kandoelec_TP.to_excel(writer, sheet_name = 'KANDOELECTPs', index = False)
# data_Ekepd_SR.to_excel(writer, sheet_name = 'EKEDPSRs', index = False)
# data_Ekedp_TP.to_excel(writer, sheet_name = 'EKEDPTPs', index = False)
# data_Jedc_SR.to_excel(writer, sheet_name = 'JEDCSRs', index = False)
# data_Jedc_TP.to_excel(writer, sheet_name = 'JEDCTPs', index = False)

# writer.save()

# data_k2 = pd.read_csv('tab_5.csv')

# #.......Mobiles Report

# data_k3 = (data_k2['Operator'] == 'AIRT') | (data_k2['Operator'] == 'ETST') | (data_k2['Operator'] == 'GLO') | (data_k2['Operator'] == 'MTN')
# data_k3 = data_k2.loc[data_k3]
# pivot = data_k3.pivot_table('Amount',['Client Type'],columns=['Operator'],aggfunc = np.sum, margins = True, margins_name = 'Grand Total')
# # To put into workbook/worksheet 
# writer_Telcos = pd.ExcelWriter('20221206 OneCard Telco Utilization Report Nov.,2022.xlsx')
# data_k3.to_excel(writer_Telcos, sheet_name = 'Telcos_Report', index = False)
# pivot.to_excel(writer_Telcos, sheet_name = 'Telcos_Summary',startrow = 1)

# writer_Telcos.save()

# # To design
# wb = load_workbook('20221206 OneCard Telco Utilization Report No  v.,2022.xlsx')
# ws = wb['Telcos_Summary']
# ws1 = wb['Telcos_Report']
# ws.sheet_properties.tabColor = '000000FF'
# ws1.sheet_properties.tabColor = '228B22'
# ws.merge_cells('A1:F1')
# ws['A1'] = 'Ade'
# centa = ws['A1']
# centa.alignment = Alignment(horizontal='center')                  #ws2['B'] = ws2['B'].astypes(float)

# thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
#                      right=Side(border_style='dashed',color='FF000000'),
#                      top=Side(border_style='thin',color='FF000000'),
#                      bottom=Side(border_style='thin',color='FF000000')
#                     )
# thick_border = Border(left=Side(border_style='thin',color='FF000000'),
#                      right=Side(border_style='thin',color='FF000000'),
#                      top=Side(border_style='thin',color='FF000000'),
#                      bottom=Side(border_style='medium',color='FF000000')
#                     )
# thin_border = Border(left=Side(border_style='thin',color='FF000000'),
#                      right=Side(border_style='thin',color='FF000000'),
#                      top=Side(border_style='thin',color='FF000000'),
#                      bottom=Side(border_style='thin',color='FF000000')
#                     )

# fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='0099CCFF',end_color='0099CCFF')
# fill1_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='9ACD32',end_color='9ACD32')

# row_num = 8
# col_num = 7
# row_loc=1
# col_loc=1

# for i in range (row_loc,row_loc*row_num):
#     for j in range (col_loc,col_loc*col_num):
#         ws.cell(row=i*1,column=j*1).border=thin_border
#         if i==row_loc:
#             ws.cell(row=i+0,column=j+0).border=thin_border
#             ws.cell(row=i+1,column=j+0).fill=fill_cell
#             ws.cell(row=i+0,column=j+0).border=thin_border
#             ws.cell(row=i+6,column=j+0).fill=fill_cell
#             ws.cell(row=i+0,column=j+0).fill=fill1_cell

#         if i==row_loc*row_num+1:
#             ws.cell(row=i+0,column=j+0).border=thick_border

# wb.save('20221206 OneCard Telco Utilization Report Nov.,2022.xlsx')

#.......ARTEE REPORT

# data_k2 = pd.read_csv('tab_5.csv')

# data_k4 = (data_k2['Agent Code'] == 'TPR_ARTEE')|(data_k2['Agent Code'] == 'TPR_Artee_CAL')|(data_k2['Agent Code'] == 'TPR_Artee_Ikoyi')|(data_k2['Agent Code'] == 'TPR_Artee_Enug')|(data_k2['Agent Code'] == 'TPR_Artee_Lekki')|(data_k2['Agent Code'] == 'TPR_Artee_Wuse')|(data_k2['Agent Code'] == 'TPR_Artee_Opebi') | (data_k2['Agent Code'] == 'TPR_Artee_Mall')|(data_k2['Agent Code'] == 'TPR_Artee_Gson')|(data_k2['Agent Code'] == 'TPR_Artee_Surul')|(data_k2['Agent Code'] == 'TPR_Artee_Vic')|(data_k2['Agent Code'] == 'TPR_Artee_Ikeja')|(data_k2['Agent Code'] == 'TPR_Artee_CEDI')|(data_k2['Agent Code'] == 'TPR_Artee_MMIA') # To bring out Artee
# data_k4 = data_k2.loc[data_k4]
# pivot = data_k4.pivot_table('Amount',['Agent Code'],columns=['Operator'],aggfunc = np.sum, margins = True, margins_name = 'Grand Total')
# pivot.insert(1,'A_Commission', '')
# pivot['A_Commission'] = [3.5/100] * pivot['AIRT']
# pivot.insert(3,'E_Commission', '')
# pivot['E_Commission'] = [4/100] * pivot['ETST']
# pivot.fillna(0,inplace=True)
# pivot.insert(5,'G_Commission', '')
# pivot['G_Commission'] = [3/100] * pivot['GLO']
# pivot.insert(7,'M_Commission', '')
# pivot['M_Commission'] = [2.5/100] * pivot['MTN']
# pivot.insert(9,'G_Total_Commission', '')
# pivot['G_Total_Commission'] = pivot['A_Commission'] + pivot['E_Commission'] + pivot['G_Commission'] + pivot['M_Commission']
# writer_Artee = pd.ExcelWriter('20221206 OneCard Spar Monthly Utilization-Commission Nov.,2022.xlsx')
# data_k4.to_excel(writer_Artee, sheet_name = 'Artee_Report', index = False)
# pivot.to_excel(writer_Artee, sheet_name = 'Artee_Commission_Summary',startrow=1)

# writer_Artee.save()

# wb = load_workbook('20221206 OneCard Spar Monthly Utilization-Commission Nov.,2022.xlsx')
# ws = wb['Artee_Commission_Summary']
# ws1 = wb['Artee_Report']
# ws.sheet_properties.tabColor = '000000FF'
# ws1.sheet_properties.tabColor = '228B22'
# ws.merge_cells('A1:K1')
# ws['A1'] = 'Ade'
# centa = ws['A1']
# centa.alignment = Alignment(horizontal='center')

# thin_border = Border(left=Side(border_style='dashed',color='FF000000'),
#                      right=Side(border_style='dashed',color='FF000000'),
#                      top=Side(border_style='thin',color='FF000000'),
#                      bottom=Side(border_style='thin',color='FF000000')
#                     )
# thick_border = Border(left=Side(border_style='thin',color='FF000000'),
#                      right=Side(border_style='thin',color='FF000000'),
#                      top=Side(border_style='thin',color='FF000000'),
#                      bottom=Side(border_style='medium',color='FF000000')
#                     )
# thin_border = Border(left=Side(border_style='thin',color='FF000000'),
#                      right=Side(border_style='thin',color='FF000000'),
#                      top=Side(border_style='thin',color='FF000000'),
#                      bottom=Side(border_style='thin',color='FF000000')
#                     )

# fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='0099CCFF',end_color='0099CCFF')
# fill1_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='9ACD32',end_color='9ACD32')

# row_num = 17
# col_num = 12
# row_loc=1
# col_loc=1

# for i in range (row_loc,row_loc*row_num):
#     for j in range (col_loc,col_loc*col_num):
#         ws.cell(row=i*1,column=j*1).border=thin_border
#         if i==row_loc:
#             ws.cell(row=i+0,column=j+0).border=thin_border
#             ws.cell(row=i+1,column=j+0).fill=fill_cell
#             ws.cell(row=i+0,column=j+0).border=thin_border
#             ws.cell(row=i+15,column=j+0).fill=fill_cell
#             ws.cell(row=i+0,column=j+0).fill=fill1_cell
            
#         if i==row_loc*row_num+1:
#             ws.cell(row=i+0,column=j+0).border=thick_border

# wb.save('20221206 OneCard Spar Monthly Utilization-Commission Nov.,2022.xlsx')


