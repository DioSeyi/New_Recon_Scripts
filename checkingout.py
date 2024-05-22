import os
import pandas as pd
import numpy as np
import datetime as dt
from glob import glob
import xlsxwriter
from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment,fills
pd.set_option("display.max_rows", 5)

tab_1 = glob('transactiondetail_*.csv')
tab_2 = pd.concat([pd.read_csv(f) for f in tab_1], sort = False) # List comprehension
tab_2.drop(['Transaction Detail','Unnamed: 3','Unnamed: 4','Unnamed: 9'],axis=1,inplace=True)
tab_2.columns = tab_2.iloc[2]
tab_2.drop([0,1,2],inplace = True)
tab_9 = tab_2['Result Description'].str.contains('Transaction Successful',na = False)|tab_2['Result Description'].str.contains('InProgress',na = False)
tab_9 = tab_2.loc[tab_9]
tab_10 = (tab_9['Trans Type'] == 'Topup') | (tab_9['Trans Type'] == 'Fund Transfer') | (tab_9['Trans Type'] == 'Wallet Adjustment') | (tab_9['Trans Type'] == 'Wallet Transfer') | (tab_9['Trans Type'] == 'Wallet Topup') | (tab_9['Trans Type'] == 'PINREDEEM') # To disregards any other top inclusive like 'Wallet Topup'
tab_10 = tab_9.loc[tab_10]
tab_10['Client Type'].fillna('B2B',inplace=True)
tab_10.to_csv('tab_10.csv')
