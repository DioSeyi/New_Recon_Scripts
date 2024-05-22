#!/usr/bin/env python
# coding: utf-8

# In[6]:


cd C:\Users\IT DEPARTMENT\Desktop\pydata


# In[7]:


pwd!


# In[8]:


import os
import pandas as pd
import numpy as np
import datetime as dt
from glob import glob
import xlsxwriter
from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill,Border,Side,Alignment,fills
pd.set_option("display.max_rows", 5)


# In[9]:


check = pd.read_excel('Ibu__Sha.xlsx')


# In[20]:


check.columns


# In[11]:


check.rename({'Unnamed: 10':'Ibu','Unnamed: 11':'Sha'},axis = 1,inplace = True)


# In[ ]:





# In[13]:


check.count()


# In[22]:


tab_9 = check['Ibu'].str.contains('Ibucap',na = False)
tab_9 = check.loc[tab_9]
tab_9.drop(['Sha'],axis=1,inplace=True)
tab_9.columns = tab_9.iloc[2]
tab_9.to_excel('tab_9.xlsx')
tab_10 = check['Sha'].str.contains('Shalarthem',na = False)
tab_10 = check.loc[tab_10]
tab_10.drop(['Ibu'],axis=1,inplace=True)
tab_10.columns = tab_10.iloc[2]
tab_10.to_excel('tab_10.xlsx')


# In[15]:


divid = pd.read_csv('switch.csv')  # pd.read_csv('switch.csv')


# In[133]:


divid.head(5)


# In[134]:


divid['SerialNo'] = divid['SerialNo'].str.replace(r'\W',"")


# In[135]:


divid.head(5)


# In[136]:


for column in divid.columns:
     divid[column] = divid[column].astype(str).str.replace(r'\W',"")


# In[137]:


divid.head(5)


# In[138]:


divid.head(5)


# In[140]:


divid['SerialNo'] = divid['SerialNo'].str.replace(r'PD',"")


# In[141]:


divid.head(5)


# In[142]:


divid['Status'] = divid['Status'].str.replace(r'_',"")


# In[143]:


divid.head(5)


# In[16]:


shalina = pd.read_csv('shalina1.csv')


# In[18]:


shalina['SerialNo'] = shalina['SerialNo'].str.replace(r'PD',"")


# In[19]:


shalina.head(5)


# In[14]:


shalina.columns


# In[ ]:


tab_1 = glob('transactiondetail_*.csv')
tab_2 = pd.concat([pd.read_csv(f) for f in tab_1], sort = False) # List comprehension
tab_2.drop(['Transaction Detail','Unnamed: 3','Unnamed: 4','Unnamed: 9'],axis=1,inplace=True)
tab_2.columns = tab_2.iloc[2]
tab_2.drop([0,1,2],inplace = True)
tab_9 = tab_2['Result Description'].str.contains('Transaction Successful',na = False)|tab_2['Result Description'].str.contains('InProgress',na = False)
tab_9 = tab_2.loc[tab_9]
tab_10 = (tab_9['Trans Type'] == 'Topup') | (tab_9['Trans Type'] == 'Fund Transfer') | (tab_9['Trans Type'] == 'Wallet Adjustment') | (tab_9['Trans Type'] == 'Wallet Transfer') | (tab_9['Trans Type'] == 'Wallet Topup') | (tab_9['Trans Type'] == 'PINREDEEM') # To disregards any other top inclusive like 'Wallet Topup'
tab_10 = tab_9.loc[tab_10]
tab_10['Client Type'].fillna('B2B',inplace=True)
tab_10.to_csv('tab_10.csv')


# In[26]:


shalina1 = shalina['Status'].str.contains('Success',na = False)
shalina1 = shalina.loc[shalina1]


# In[27]:


shalina1.head(5)


# In[29]:


shalina2 = (shalina1['Amount'] == 2000) # To disregards any other top inclusive like 'Wallet Topup'
shalina2 = shalina1.loc[shalina2]
shalina2.to_csv('shalina9.csv',index = False)


# In[25]:


# shalina = glob('Oct_sha*.csv')
# shalina = pd.concat([pd.read_csv(f) for f in tab_1], sort = False)

shalina = pd.read_csv('Oct_sha.csv')


# In[28]:


shalina['SerialNo'] = shalina['SerialNo'].str.replace(r'PD',"")


# In[29]:


shalina1 = shalina['Status'].str.contains('Success',na = False)|shalina['Status'].str.contains('InProcess',na = False)
shalina1 = shalina.loc[shalina1]


# In[30]:


shalina2 = (shalina1['Amount'] == 2000)
shalina2 = shalina1.loc[shalina2]


# In[31]:


shalina2.columns


# In[32]:


# shalina2.drop(['Transaction Detail','Unnamed: 3','Unnamed: 4'],axis=1,inplace=True)
# tab_3.columns = tab_3.iloc[2]

shalina2.to_csv('SHALINA_again.csv',index = False)


# In[30]:


import openpyxl as xl
import pprint
print('enter document name...')
docname = input('enter document name with extension:')
print('Opening workbook...')
file1 = xl.load_workbook(docname)
sheet1title = input ('enter sheet 1 title: ')
sheet1 = file1[sheet1title]
sheet2title = input ('enter sheet 2 title: ')
sheet2 = file1[sheet2title]
sheet3title = input ('enter sheet 3 title: ')
sheet3 = file1[sheet3title]
print('reading workbook')


# In[ ]:





# In[32]:


sr_sales = []
print('Reading rows...')

for row in range(2, sheet2.max_row + 1):
    agent = sheet2['A' +str(row)].value
    funding = sheet2['I' + str(row)].value
    sr_sales.append(funding)
    
    
sr1_sales = []
for row1 in range(2, sheet3.max_row + 1):
    agent = sheet3['A' +str(row1)].value
    funding = sheet3['I' + str(row1)].value
    sr1_sales.append(funding)

    #sr_sales[agent]['funding'] = funding
    #See how to roll out another comparison sheet
    
print('Reading rows...')
for rowNum in range(2, sheet1.max_row + 1):
    agent_a1 = sheet1.cell(row=rowNum, column=9).value
    if agent_a1 in sr_sales:
        sheet1.cell(row=rowNum, column = 11).value = 'Ibucap'

for rowNum1 in range(2, sheet1.max_row + 1):
    agent_a2 = sheet1.cell(row = rowNum1, column=9).value
    if agent_a2 in sr1_sales:
        sheet1.cell(row =rowNum1, column = 12).value = 'Shalarthem'
        
print ('finishing up...saving...') 
file1.save('Ibu__Sha.xlsx')
file1.close()


# In[ ]:




