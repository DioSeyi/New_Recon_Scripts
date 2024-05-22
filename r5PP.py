import openpyxl as xl
import pprint
import pandas as pd
import numpy as np


file=input('Enter a file: ')  
print('Evoking the file....')
file1=xl.load_workbook(file)
print('loading the file....')

sheet1title=input ('enter sheet 1 title: ') 
sheet1=file1[sheet1title]
print('reading workbook......')

sr_sales_A_S = []
for row in range(2, sheet1.max_row + 1): 
    Meter_Acount_Position = sheet1['E' + str(row)].value
    sr_sales_A_S.append(Meter_Acount_Position)
    
print('reading rows...')
 
for rowNum1 in range(2, sheet1.max_row + 1): 
    Meter_Acount_Number_Common = sheet1.cell(row=rowNum1, column = 19).value #& rowNum1['CUSTOMER TYPE'].str.contain('OFFLINE_POSTPAID',na = False)
    if Meter_Acount_Number_Common in sr_sales_A_S :
        sheet1.cell(row=rowNum1, column = 14).value = 'yes'

for rowNum2 in range(2, sheet1.max_row + 1): 
    Meter_Acount_Number_Common = sheet1.cell(row=rowNum2, column = 20).value
    if Meter_Acount_Number_Common in sr_sales_A_S:
        sheet1.cell(row=rowNum2, column = 15).value = 'yes' 
    
file1.save('Ok.xlsx')

file1 = pd.read_excel('Ok.xlsx')
writer_Recon = pd.ExcelWriter('Ok.xlsx')
file2 = (file1['CUSTOMER TYPE'] == 'OFFLINE_POSTPAID') #or  # file2 = file1['CUSTOMER TYPE'].str.contains('OFFLINE_PREPAID',na = True)
file3 = file1.loc[file2]
file4 = (file1['CUSTOMER TYPE'] == 'OFFLINE_PREPAID') #or  # file2 = file1['CUSTOMER TYPE'].str.contains('OFFLINE_PREPAID',na = True)
file5 = file1.loc[file4]
file1.to_excel(writer_Recon, sheet_name = 'overall_Report', index = False)
file3.to_excel(writer_Recon, sheet_name = 'Postpaid_Recon_Report', index = False)
file5.to_excel(writer_Recon, sheet_name = 'Prepaid_Recon_Report',index=False)

writer_Recon.save()